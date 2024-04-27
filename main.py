import os
import openpyxl
import requests
import time
from json import loads
from bs4 import BeautifulSoup

page_url = 'https://shop.adidas.jp/men/'


class ExcelWriter:
    headers = {}
    def __init__(self, filename):
        self.filename = filename
        if os.path.exists(filename):
            self.workbook = openpyxl.load_workbook(filename)
        else:
            self.workbook = openpyxl.Workbook()
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
        self.worksheets = {ws.title: ws for ws in self.workbook.worksheets}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.save()

    def add_worksheet(self, title):
        if title not in self.worksheets:
            worksheet = self.workbook.create_sheet(title=title)
            self.worksheets[title] = worksheet
            ExcelWriter.headers[title] = False
        else:
            worksheet = self.worksheets[title]
        return worksheet

    def write_data(self, worksheet, data):

        if data['reviews'] is not '':
            data['reviews'] = '\n'.join([str(d) for d in data['reviews']])

        if not ExcelWriter.headers[worksheet]:
            ExcelWriter.headers[worksheet] = True
            headers = list(data.keys()) if isinstance(data, dict) else list(data[0].keys())
            self.worksheets[worksheet].append(headers)
        if isinstance(data, dict):
            self.worksheets[worksheet].append([data[header] for header in data.keys()])
        else:
            for row in data:
                self.worksheets[worksheet].append([row[header] for header in row.keys()])

    def save(self):
        self.workbook.save(self.filename)


def get_details(content, tag, selector):
    try:
        soup = BeautifulSoup(content.text, 'html.parser')
        text = soup.find(tag, class_=selector)
        return text.text
    except Exception:
        return ""


def get_available_size(content, tag, selector):
    try:
        soup = BeautifulSoup(content.text, 'html.parser')
        size = soup.find(tag, class_=selector)
        all_sizes = ' '.join(size.stripped_strings)
        return all_sizes
    except Exception:
        return ""


def get_breadcrumbs(content, tag, selector):
    try:
        soup = BeautifulSoup(content.text, 'html.parser')
        bread = soup.find(tag, class_=selector)
        breadcrumbs = [item.text.strip() for item in bread.find_all("li")]
        return ' '.join(breadcrumbs)
    except Exception:
        return ""


def get_sense(content, tag, selector):
    try:
        soup = BeautifulSoup(content.text, 'html.parser')
        sense = soup.find(tag, class_=selector)
        if not sense:
            return None
        else:
            sense = sense.get('class')
        record = 0
        for sc in sense:
            if sc.startswith('mod-marker_'):
                record = sc.strip('mod-marker_').replace('_', '.')
                break
        return record
    except Exception:
        return ""


def get_list_of_items(content, tag, classes, single=False):
    if not content:
        return ''
    soup = content
    if isinstance(content, str):
        soup = BeautifulSoup(content, 'html.parser')

    finder = soup.find if single else soup.findAll
    return finder(tag, classes)


def get_attribute(soup_item, attribute):
    if not soup_item:
        return None
    return soup_item.get(attribute)


def get_product_default_images(content, as_json=''):
    images = get_list_of_items(content, 'img', {'class': 'test-image'})
    images = [get_attribute(image, 'src') for image in images if image]
    fmt = 'https://shop.adidas.jp{}'
    loader_images = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('article', {})
        .get('image', {})
        .get('details', [])
    )
    loader_images = [li.get('imageUrl', {}).get('large', '') for li in loader_images if li]
    images = [fmt.format(img) for img in (images + loader_images) if img and 'itemCard_dummy.jpg' not in img]
    return ','.join(images)


def additional_review_data(code, model=None):
    url = 'https://adidasjp.ugc.bazaarvoice.com/7896-ja_jp/{}/reviews.djs'.format(model)
    data = {'format': 'embeddedhtml', 'productattribute_itemKcod': code, 'page': 1}
    response = requests.get(url, data).content.decode().split('\n')
    review_bar = response[6].replace('var materials=', '').replace('},', '').replace('\n', '').replace('\\', '').strip('n"')
    review = get_list_of_items(review_bar, 'span', {'class': 'BVRRNumber', 'itemprop': 'ratingValue'}, True)
    review = review.text if review else ''
    percentage = get_list_of_items(review_bar, 'span', {'class': 'BVRRBuyAgainPercentage'}, True)
    percentage = percentage.text if percentage else ''

    container = get_list_of_items(review_bar, 'div', {'class': 'BVRRSecondaryRatingsContainer'}, True)
    if not container:
        return review, percentage, '', '', '', ''
    fit = get_list_of_items(container, 'div', {'class': 'BVRRRatingFit'}, True)
    fit = get_list_of_items(fit, 'img', {'class': 'BVImgOrSprite'}, True)
    fit = get_attribute(fit, 'alt') if fit else ''

    length = get_list_of_items(container, 'div', {'class': 'BVRRRatingLength'}, True)
    length = get_list_of_items(length, 'img', {'class': 'BVImgOrSprite'}, True)
    length = get_attribute(length, 'alt') if length else ''

    quality = get_list_of_items(container, 'div', {'class': 'BVRRRatingQuality'}, True)
    quality = get_list_of_items(quality, 'img', {'class': 'BVImgOrSprite'}, True)
    quality = get_attribute(quality, 'alt') if quality else ''

    comfort = get_list_of_items(container, 'div', {'class': 'BVRRRatingComfort'}, True)
    comfort = get_list_of_items(comfort, 'img', {'class': 'BVImgOrSprite'}, True)
    comfort = get_attribute(comfort, 'alt') if comfort else ''
    return review, percentage, fit, length, quality, comfort


def get_product_reviews(as_json):
    reviews = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('model', {})
        .get('review', {})
    )
    code = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('productIdInQuery', '')
    )
    
    model = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('article', {})
        .get('modelCode', '')
    )
    avg_rev, percentage, fit, length, quality, comfort = additional_review_data(code, model)
    reviews['avg_rev'] = avg_rev
    reviews['percentage'] = percentage
    reviews['fit'] = fit
    reviews['length'] = length
    reviews['quality'] = quality
    reviews['comfort'] = comfort
    return reviews


def build_reviews(reviews, related_for):
    if not reviews:
        return ''
    for review in reviews:
        review['author'] = review['author']['name']
        review['reviewRating'] = review['reviewRating']['ratingValue']
        # review['related_for'] = related_for
    return reviews


def men_category():
    resp = requests.get(page_url)
    soup = BeautifulSoup(resp.text, 'html.parser')
    elements = soup.find_all('a', class_='lpc-teaserCarousel_link')
    href_list = [element['href'] for element in elements]
    return href_list


def product_details(product_urls):
    for product_url in product_urls:
        time.sleep(0.3)
        content = requests.get(product_url)
        try:
            content_decode = content.content.decode()
        except Exception:
            content_decode = ''
        loader_data = get_list_of_items(content_decode, 'script', {'id': '__NEXT_DATA__'}, True)
        as_json = loads(loader_data.text) if loader_data else {}
        reviews = get_product_reviews(as_json)
        details = {
            'url': product_url,
            'breadcrumb': get_breadcrumbs(content, 'ul', 'breadcrumbList'),
            'category': get_details(content, 'a', 'groupName'),
            'images': get_product_default_images(content_decode, as_json),
            'product_name': get_details(content, 'h1', 'itemTitle'),
            'price': get_details(content, 'span', 'price-value'),
            'available_sizes': get_available_size(content, 'ul', 'sizeSelectorList'),
            'sense': get_sense(content, 'span', 'test-marker'),
            'title_description': get_details(content, 'h4', 'itemFeature'),
            'general_description': get_details(content, 'div', 'commentItem-mainText'),
            'review_count': reviews['reviewCount'],
            'avg_rev': reviews['avg_rev'],
            'percentage': reviews['percentage'],
            'fit': reviews['fit'],
            'length': reviews['length'],
            'quality': reviews['quality'],
            'comfort': reviews['comfort'],
            'reviews': build_reviews(reviews['reviewSeoLd'], product_url),
        }
        print(details)
        with ExcelWriter(filename='addidas.xlsx') as ew:
            ew.write_data('details', details)
        break


def items(category):
    params = category.split('?')[-1] + '&limit=40'
    url = 'https://shop.adidas.jp/f/v1/pub/product/list?' + params
    resp = requests.get(url)
    if not resp.ok:
        return {}
    resp = resp.json()
    list_of_items = resp['articles_sort_list']
    time.sleep(2)
    product_urls = [
        f'https://shop.adidas.jp/products/{item}/' for item in list_of_items
    ]
    print(product_urls)
    return product_details(product_urls)


def main():
    categories = men_category()
    with ExcelWriter(filename='addidas.xlsx') as ew:
        ew.add_worksheet('details')
    for category in categories:
        items(category)


main()
